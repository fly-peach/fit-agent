"""种子脚本：从 Excel 导入食物数据到数据库。

用法：
    cd rogers
    python scripts/seed_foods.py
"""
import sys
import openpyxl

EXCEL_PATH = r"C:\Users\wt197\Downloads\热量数据大全_v4_优化版.xlsx"


def main():
    print("正在读取 Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

    # Sheet 1: 原材料
    ws1 = wb.worksheets[0]
    rows1 = list(ws1.iter_rows(values_only=True))
    header1 = rows1[0]
    print(f"  Sheet1: {header1}")
    print(f"  {len(rows1) - 1} 条原材料数据")

    # Sheet 2: 菜品
    ws2 = wb.worksheets[1]
    rows2 = list(ws2.iter_rows(values_only=True))
    header2 = rows2[0]
    print(f"  Sheet2: {header2}")
    print(f"  {len(rows2) - 1} 条菜品数据")

    # Generate SQL
    sql_lines = []
    sql_lines.append("-- 自动生成的食物数据库种子脚本")
    sql_lines.append("-- 由 seed_foods.py 生成")
    sql_lines.append("")
    sql_lines.append("DELETE FROM food_items WHERE source = 'system';")
    sql_lines.append("")

    total = 0

    # 原材料
    for row in rows1[1:]:
        name, category, portion_unit, portion_g, portion_cal, cal_100g, cal_level = row
        sql_lines.append(
            f"INSERT INTO food_items (name, category, source, portion_unit, portion_grams, "
            f"portion_calories, calories_per_100g, calorie_level) VALUES "
            f"('{str(name).replace(chr(39), chr(39)+chr(39))}', '{str(category)}', 'system', "
            f"'{str(portion_unit)}', {int(portion_g)}, {int(portion_cal)}, {int(cal_100g)}, '{str(cal_level)}');"
        )
        total += 1

    # 菜品
    for row in rows2[1:]:
        name, category, portion_unit, portion_g, portion_cal, cal_100g, cal_level = row
        sql_lines.append(
            f"INSERT INTO food_items (name, category, source, portion_unit, portion_grams, "
            f"portion_calories, calories_per_100g, calorie_level) VALUES "
            f"('{str(name).replace(chr(39), chr(39)+chr(39))}', '{str(category)}', 'system', "
            f"'{str(portion_unit)}', {int(portion_g)}, {int(portion_cal)}, {int(cal_100g)}, '{str(cal_level)}');"
        )
        total += 1

    sql_lines.append(f"\n-- 共插入 {total} 条食物数据")

    output_path = __file__.replace("seed_foods.py", "seed_foods.sql")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(sql_lines))

    print(f"\n 已生成 {output_path}（{total} 条 INSERT 语句）")
    print("运行方式：")
    print(f"  sqlite3 db/fitagent.db < {output_path}")


if __name__ == "__main__":
    main()
